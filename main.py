from pathlib import Path

import openpyxl
import xlsxwriter

loc = "respuestas.xlsx"

# Load the workbook and select the first sheet
try:
    wb = openpyxl.load_workbook(loc)
except FileNotFoundError:
    print(f"File {loc} not found.")
    exit(1)
except openpyxl.utils.exceptions.InvalidFileException as e:
    print(f"Error reading {loc}: {e}")
    exit(1)

sheet = wb.active

# Read header
header = [cell.value for cell in sheet[1]]

for row in sheet.iter_rows(min_row=2, values_only=True):
    nombre = row[1]
    secc = row[2][:9]

    # Define the output directory and ensure it exists
    output_dir = Path("estudiantes") / secc
    output_dir.mkdir(parents=True, exist_ok=True)

    # Define the file path
    file_path = output_dir / f"{nombre}.xlsx"

    # Create a new workbook with xlsxwriter
    wb_out = xlsxwriter.Workbook(str(file_path))
    sheet1 = wb_out.add_worksheet("Respuestas")

    # Write header
    for col_idx in range(1, len(header)):
        sheet1.write(col_idx - 1, 0, header[col_idx])

    # Write response
    for col_idx in range(1, len(row)):
        sheet1.write(col_idx - 1, 1, row[col_idx])

    # Formatting
    sheet1.set_column("A:A", 100)
    sheet1.set_column("B:B", 20)
    sheet1.conditional_format(
        2,
        1,
        32,
        1,  # Corresponds to range B3:B33 (0-indexed)
        {
            "type": "3_color_scale",
            "min_value": 1,
            "max_value": 3,
            "min_color": "#63BE7B",
            "mid_color": "#FFEB84",
            "max_color": "#F8696B",
        },
    )

    wb_out.close()
